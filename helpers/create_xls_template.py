import pandas as pd
import logging
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from helpers.metadata_check import get_columns_data

logger = logging.getLogger("my_app_logger")


def create_template_one_drive_and_excel():

    metadata = get_columns_data(exclude=True)
    logger.debug(f"Metadata: {metadata}")

    # Create a DataFrame with columns based on the metadata keys
    columns = list(metadata.keys())
    df = pd.DataFrame(columns=columns)

    # Create a new Excel workbook and select the active worksheet
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Template"

    # Define the columns for the main sheet
    ws_data.append(columns)

    # Create a secondary sheet for options
    ws_options = wb.create_sheet(title="Options")

    # Iterate through the metadata to find columns with 'options'
    for col, details in metadata.items():
        if "options" in details:
            category_options = details["options"]
            col_idx = df.columns.get_loc(col) + 1  # Get column index (1-based)
            col_letter = get_column_letter(col_idx)
            # Write options to the options sheet
            for i, option in enumerate(category_options, start=1):
                ws_options.cell(row=i, column=col_idx, value=option)

            # Directly reference the options range in the DataValidation
            option_range = (
                f"Options!${col_letter}$1:"
                f"${col_letter}${len(category_options)}"
            )
            dv_category = DataValidation(
                type="list", formula1=f"={option_range}", showDropDown=False
            )

            # Add data validation to the 'Category' column
            ws_data.add_data_validation(dv_category)

            # Apply data validation to each cell in the column
            # first 100 rows for example)
            for row in range(2, 102):
                dv_category.add(ws_data.cell(row=row, column=col_idx))

    # Save the workbook
    file_path = "template_with_dropdowns_for_one_drive_and_excel.xlsx"
    try:
        wb.save(file_path)
        logger.info(f"Simple template with dropdown saved as {file_path}")
    except Exception as e:
        logger.error(f"Failed to save the workbook: {e}")
